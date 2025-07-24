import sys
import os
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
from conexion2 import conectarsql

def generar_reporte_excel():
    db_connection = conectarsql()
        
    with db_connection.cursor() as cursor:
        query = """
        SELECT 
            a.Fecha_Log, 
            u2.Usuario AS [usuario modificador], 
            a.Accion,
            u1.Usuario AS [usuario modificado],
            R.Cod_Tienda
        FROM AuditLog a
        JOIN Users u1 ON a.Cod_Usuario = u1.Cod_Usuario 
        JOIN Users u2 ON a.Modificador = u2.Cod_Usuario
        LEFT JOIN Restaurante R ON a.Cod_Restaurante = R.Cod_Restaurante
        ORDER BY a.Fecha_Log DESC
        """

        cursor.execute(query)
        rows = cursor.fetchall()

        if not rows:
            return None
    cursor.close()
    
    # Crear archivo Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte Acciones realizadas por el Bot"

    # Encabezados
    headers = ["Fecha_Log", "Usuario que ingreso", "Acción", "Usuario Modificado", "Código de tienda"]
    ws.append(headers)

    #Estilos de la tabla
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center')
    border_style = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = bold_font
        cell.alignment = center_align
        cell.border = border_style
    
    # Agregar los datos
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, value in enumerate(list(row), start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border_style
            
    for col in ws.columns:
        max_length = 0
        column = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = max_length + 2

    # Guardar el archivo
    output_path = "reporte_Bot.xlsx"
    wb.save(output_path)
    return output_path
    

if __name__ == "__main__":
    ruta = generar_reporte_excel()
    if ruta:
        print(f"✅ Reporte generado correctamente: {ruta}")
    else:
        print("⚠️ No se generó el reporte (posiblemente no hay datos).")
    